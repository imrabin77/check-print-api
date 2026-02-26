import csv
import io
import os
import uuid
from datetime import date
from decimal import Decimal, InvalidOperation
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from sqlalchemy.orm import joinedload
from app.core.database import get_db
from app.core.security import allow_all_roles, allow_admin_clerk, allow_admin
from app.models.invoice import Invoice
from app.models.vendor import Vendor
from app.models.check import Check
from app.models.user import User
from app.schemas.invoice import InvoiceResponse, InvoiceUpdate, ImportSummary, InvoiceCreate

router = APIRouter(prefix="/api/invoices", tags=["invoices"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _to_response(inv: Invoice) -> InvoiceResponse:
    return InvoiceResponse(
        id=inv.id,
        invoice_number=inv.invoice_number,
        store_number=inv.store_number,
        vendor_id=inv.vendor_id,
        vendor_name=inv.vendor.name if inv.vendor else None,
        amount=inv.amount,
        invoice_date=inv.invoice_date,
        status=inv.status,
        check_id=inv.check_id,
        check_number=inv.check.check_number if inv.check else None,
        notes=inv.notes,
        attachment_filename=inv.attachment_filename,
        source_type=inv.source_type,
        imported_by_name=inv.imported_by.full_name if inv.imported_by else None,
        imported_at=inv.imported_at,
        created_at=inv.created_at,
    )


@router.get("", response_model=list[InvoiceResponse])
async def list_invoices(
    status: str | None = Query(None),
    search: str | None = Query(None),
    store: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_all_roles),
):
    query = select(Invoice).options(joinedload(Invoice.vendor), joinedload(Invoice.check), joinedload(Invoice.imported_by))
    if status:
        query = query.where(Invoice.status == status)
    if store:
        query = query.where(Invoice.store_number == store)
    if search:
        query = query.join(Invoice.vendor).where(
            or_(
                Invoice.invoice_number.ilike(f"%{search}%"),
                Vendor.name.ilike(f"%{search}%"),
                Invoice.store_number.ilike(f"%{search}%"),
            )
        )
    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    return [_to_response(inv) for inv in result.unique().scalars().all()]


@router.get("/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_all_roles),
):
    result = await db.execute(
        select(Invoice)
        .options(joinedload(Invoice.vendor), joinedload(Invoice.check), joinedload(Invoice.imported_by))
        .where(Invoice.id == invoice_id)
    )
    inv = result.unique().scalar_one_or_none()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return _to_response(inv)


@router.put("/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(
    invoice_id: int,
    body: InvoiceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_all_roles),
):
    result = await db.execute(
        select(Invoice).options(joinedload(Invoice.vendor), joinedload(Invoice.check), joinedload(Invoice.imported_by)).where(Invoice.id == invoice_id)
    )
    inv = result.unique().scalar_one_or_none()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    for key, value in body.model_dump(exclude_unset=True).items():
        setattr(inv, key, value)
    await db.flush()
    await db.refresh(inv)
    return _to_response(inv)


@router.post("/{invoice_id}/approve", response_model=InvoiceResponse)
async def approve_invoice(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_admin),
):
    result = await db.execute(
        select(Invoice).options(joinedload(Invoice.vendor), joinedload(Invoice.check), joinedload(Invoice.imported_by)).where(Invoice.id == invoice_id)
    )
    inv = result.unique().scalar_one_or_none()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.status != "PENDING":
        raise HTTPException(status_code=400, detail=f"Cannot approve invoice with status {inv.status}")
    inv.status = "APPROVED"
    await db.flush()
    await db.refresh(inv)
    return _to_response(inv)


# ─── Bulk import: CSV + Excel ────────────────────────────────────────

async def _find_or_create_vendor(name: str, db: AsyncSession) -> Vendor:
    v_result = await db.execute(select(Vendor).where(Vendor.name == name))
    vendor = v_result.scalar_one_or_none()
    if not vendor:
        vendor = Vendor(name=name)
        db.add(vendor)
        await db.flush()
    return vendor


async def _import_rows(rows: list[dict], source: str, db: AsyncSession, user_id: int) -> ImportSummary:
    total = imported = skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        total += 1
        inv_num = str(row.get("invoice_number", "")).strip()
        store_num = str(row.get("store_number", "") or "").strip()
        vendor_name = str(row.get("vendor_name", "")).strip()
        amount_str = str(row.get("amount", "")).strip()
        date_str = str(row.get("invoice_date", "") or "").strip()

        if not inv_num or not vendor_name or not amount_str or not store_num:
            errors.append(f"Row {i}: missing required field(s)")
            continue

        try:
            amount = Decimal(amount_str)
        except InvalidOperation:
            errors.append(f"Row {i}: invalid amount '{amount_str}'")
            continue

        inv_date = date.today()
        if date_str:
            try:
                from dateutil.parser import parse as parse_date
                inv_date = parse_date(date_str).date()
            except (ValueError, TypeError):
                errors.append(f"Row {i}: invalid date '{date_str}'")
                continue

        existing = await db.execute(select(Invoice).where(Invoice.invoice_number == inv_num))
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        vendor = await _find_or_create_vendor(vendor_name, db)

        db.add(Invoice(
            invoice_number=inv_num,
            store_number=store_num,
            vendor_id=vendor.id,
            amount=amount,
            invoice_date=inv_date,
            status="PENDING",
            source_type=source,
            imported_by_id=user_id,
            notes=str(row.get("notes", "") or "").strip() or None,
        ))
        imported += 1

    await db.flush()
    return ImportSummary(total_rows=total, imported=imported, skipped_duplicates=skipped, errors=errors)


@router.post("/import", response_model=ImportSummary)
async def import_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_all_roles),
):
    filename = file.filename.lower()
    content = await file.read()

    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        required = {"invoice_number", "store_number", "vendor_name", "amount"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise HTTPException(status_code=400, detail=f"File must contain columns: {', '.join(required)}")
        rows = list(reader)
        return await _import_rows(rows, "csv", db)

    elif filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        required = {"invoice_number", "store_number", "vendor_name", "amount"}
        if not required.issubset(set(headers)):
            raise HTTPException(status_code=400, detail=f"Excel must contain columns: {', '.join(required)}")

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            rows.append(row_dict)
        wb.close()
        return await _import_rows(rows, "excel", db)

    else:
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")


# ─── Single invoice: manual entry with optional file attachment ──────

@router.post("/create", response_model=InvoiceResponse)
async def create_invoice_manual(
    invoice_number: str = Form(...),
    store_number: str = Form(...),
    vendor_id: int = Form(...),
    amount: str = Form(...),
    invoice_date: str = Form(...),
    notes: str = Form(""),
    file: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allow_all_roles),
):
    # Validate amount
    try:
        amt = Decimal(amount)
    except InvalidOperation:
        raise HTTPException(status_code=400, detail=f"Invalid amount: {amount}")

    # Parse date
    try:
        from dateutil.parser import parse as parse_date
        inv_date = parse_date(invoice_date).date()
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail=f"Invalid date: {invoice_date}")

    # Check duplicate
    existing = await db.execute(select(Invoice).where(Invoice.invoice_number == invoice_number))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Invoice {invoice_number} already exists")

    # Validate vendor exists
    vendor_result = await db.execute(select(Vendor).where(Vendor.id == vendor_id))
    vendor = vendor_result.scalar_one_or_none()
    if not vendor:
        raise HTTPException(status_code=400, detail="Vendor not found")

    # Save attachment if provided
    attachment_name = None
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        allowed = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".tiff", ".bmp"}
        if ext not in allowed:
            raise HTTPException(status_code=400, detail=f"File type {ext} not supported. Allowed: {', '.join(allowed)}")
        unique_name = f"{uuid.uuid4().hex}{ext}"
        filepath = os.path.join(UPLOAD_DIR, unique_name)
        file_content = await file.read()
        with open(filepath, "wb") as f:
            f.write(file_content)
        attachment_name = unique_name

    invoice = Invoice(
        invoice_number=invoice_number,
        store_number=store_number,
        vendor_id=vendor.id,
        amount=amt,
        invoice_date=inv_date,
        status="PENDING",
        source_type="upload" if attachment_name else "manual",
        imported_by_id=current_user.id,
        notes=notes.strip() or None,
        attachment_filename=attachment_name,
    )
    db.add(invoice)
    await db.flush()
    await db.refresh(invoice)

    # Reload with joins
    result = await db.execute(
        select(Invoice).options(joinedload(Invoice.vendor), joinedload(Invoice.check), joinedload(Invoice.imported_by)).where(Invoice.id == invoice.id)
    )
    return _to_response(result.unique().scalar_one())


# ─── Serve uploaded attachments ──────────────────────────────────────

@router.get("/attachment/{filename}")
async def get_attachment(
    filename: str,
    current_user: User = Depends(allow_all_roles),
):
    filepath = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(filepath)


# ─── OCR: Extract fields from uploaded image/PDF ────────────────────

@router.post("/ocr")
async def ocr_extract_fields(
    file: UploadFile = File(...),
    current_user: User = Depends(allow_all_roles),
):
    """Upload an image or PDF and get back extracted invoice fields."""
    allowed_ext = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".tiff", ".bmp"}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        from app.core.ocr import ocr_extract
        result = ocr_extract(content, file.filename)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR failed: {str(e)}")
